"""
build_career_kb.py — 就业知识库建库脚本
=========================================
核心思路：Excel 结构化数据 → 统计分析 → 生成文字摘要 chunk → 入库
生成的 chunk 类型：
1. 专业就业全景（每个专业一个）
2. 行业薪酬分析（每个主要行业一个）
3. 热门公司推荐（按录用人数排名）
4. 升学 vs 就业对比（全校 + 各专业）
5. 热门岗位分析
6. 年度对比摘要（多年数据时）

用法：
python build_career_kb.py build --files 就业信息下载 2025 年.xlsx
python build_career_kb.py build --files 2023.xlsx 2024.xlsx 2025.xlsx
python build_career_kb.py verify
"""

import os
import re
import argparse
import statistics
from pathlib import Path
from collections import Counter, defaultdict
from typing import List, Dict, Tuple

import openpyxl
from rich.console import Console
from rich.panel import Panel
from rich.table import Table

from career_config import CAREER_CONFIG, MAJOR_MAP
from document_processor import DocumentProcessor, Document
from hybrid_search import HybridRetriever

import hashlib

console = Console()

# ─────────────────────────────────────────────
# Excel 读取
# ─────────────────────────────────────────────

def convert_xls_to_xlsx(path: str) -> str:
    """把 .xls 文件转换成 .xlsx，返回新路径"""
    try:
        import xlrd
        from openpyxl import Workbook as openpyxlWorkbook
        new_path = path + 'x'
        rb = xlrd.open_workbook(path)
        ws_old = rb.sheet_by_index(0)
        wb_new = openpyxlWorkbook()
        ws_new = wb_new.active
        
        for row in range(ws_old.nrows):
            ws_new.append([ws_old.cell_value(row, col) for col in range(ws_old.ncols)])
        
        wb_new.save(new_path)
        console.print(f" [dim]已自动转换 {Path(path).name} → {Path(new_path).name}[/dim]")
        return new_path
    except ImportError:
        raise ImportError("请先安装 xlrd：pip install xlrd")


def load_excel(path: str, year: str = None) -> List[dict]:
    """读取就业 Excel，返回标准化记录列表，自动兼容 .xls 和 .xlsx"""
    # 自动转换旧版 .xls 格式
    if str(path).lower().endswith('.xls') and not str(path).lower().endswith('.xlsx'):
        path = convert_xls_to_xlsx(path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # 字段名映射（兼容不同年份列名略有差异的情况）
    field_map = {
        "姓名": "name",
        "就业单位名称/征兵办名称/项目名称/创业单位名称/升学院校名称/境外单位名称": "company",
        "单位类型": "unit_type",
        "单位所属行业": "industry",
        "单位/征兵办/项目/院校所属地区": "location",
        "岗位名称/专业名称/工作内容": "position",
        "职业类型": "job_type",
        "薪酬": "salary",
        "班级名称": "class_name",
        "生源地": "hometown",
    }

    col_idx = {}
    for i, h in enumerate(headers):
        if h in field_map:
            col_idx[field_map[h]] = i

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for field, idx in col_idx.items():
            rec[field] = row[idx] if row[idx] is not None else ""

        # 清理薪酬
        try:
            rec["salary"] = int(float(str(rec.get("salary", 0))))
        except (ValueError, TypeError):
            rec["salary"] = 0

        # 提取专业前缀
        cls = str(rec.get("class_name", ""))
        m = re.match(r"([^\d]+)", cls)
        rec["major_prefix"] = m.group(1) if m else "其他"
        rec["major_full"] = MAJOR_MAP.get(rec["major_prefix"], rec["major_prefix"])

        # 标注年份
        rec["year"] = year or "未知"
        records.append(rec)
    return records

# ─────────────────────────────────────────────
# 分类判断
# ─────────────────────────────────────────────
GRADUATE_KEYWORDS = {"研究生", "本科", "专升本", "升学", "考研"}
ABROAD_KEYWORDS = {"出国", "出境", "境外"}
EMPLOYMENT_EXCLUDE = {"求职中", "暂未登记", "暂未登记或上报", "待业", "灵活就业"}

def classify_record(rec: dict) -> str:
    """分类：升学 / 出国 / 就业 / 待定"""
    ut = str(rec.get("unit_type", ""))
    for kw in GRADUATE_KEYWORDS:
        if kw in ut:
            return "升学"
    for kw in ABROAD_KEYWORDS:
        if kw in ut:
            return "出国"
    for kw in EMPLOYMENT_EXCLUDE:
        if kw in ut:
            return "待定"
    return "就业"

# ─────────────────────────────────────────────
# 统计工具
# ─────────────────────────────────────────────
def salary_stats(salaries: List[int]) -> dict:
    valid = [s for s in salaries if s > 500]
    if not valid:
        return {"avg": 0, "median": 0, "min": 0, "max": 0, "count": 0}
    return {
        "avg": int(statistics.mean(valid)),
        "median": int(statistics.median(valid)),
        "min": min(valid),
        "max": max(valid),
        "count": len(valid),
    }

def top_items(counter: Counter, n: int = 5) -> List[Tuple[str, int]]:
    return [(k, v) for k, v in counter.most_common(n) if k and str(k).strip()]

# ─────────────────────────────────────────────
# Chunk 生成函数
# ─────────────────────────────────────────────

def gen_major_chunks(records: List[dict], years: List[str]) -> List[str]:
    """每个专业生成一个就业全景 chunk"""
    chunks = []
    by_major = defaultdict(list)
    for r in records:
        by_major[r["major_prefix"]].append(r)

    for prefix, recs in by_major.items():
        full_name = MAJOR_MAP.get(prefix, prefix)
        total = len(recs)
        categories = Counter(classify_record(r) for r in recs)

        employed = [r for r in recs if classify_record(r) == "就业"]
        graduate = [r for r in recs if classify_record(r) == "升学"]

        sal = salary_stats([r["salary"] for r in employed])
        top_companies = top_items(Counter(r.get("company", "") for r in employed if r.get("company")))
        top_industries = top_items(Counter(r.get("industry", "") for r in employed if r.get("industry")))
        top_positions = top_items(Counter(r.get("position", "") for r in employed if r.get("position")))
        top_locations = top_items(Counter(r.get("location", "") for r in employed if r.get("location")))
        grad_schools = top_items(Counter(r.get("company", "") for r in graduate if r.get("company")))

        employ_rate = categories.get("就业", 0) / total * 100 if total else 0
        grad_rate = categories.get("升学", 0) / total * 100 if total else 0

        lines = [
            f"# {full_name}（{prefix}）专业就业情况分析",
            f"数据年份：{'、'.join(years)}，样本总人数：{total}人",
            "",
            "## 总体去向",
            f"- 就业：{categories.get('就业', 0)}人（{employ_rate:.1f}%）",
            f"- 升学（考研/专升本）：{categories.get('升学', 0)}人（{grad_rate:.1f}%）",
            f"- 出国深造：{categories.get('出国', 0)}人",
            f"- 待定/求职中：{categories.get('待定', 0)}人",
            "",
        ]

        if sal["count"] > 0:
            lines += [
                f"## 薪酬水平（已就业 {sal['count']} 人）",
                f"- 平均薪酬：{sal['avg']} 元/月",
                f"- 中位数薪酬：{sal['median']} 元/月",
                f"- 薪酬区间：{sal['min']} ~ {sal['max']} 元/月",
                "",
            ]

        if top_companies:
            lines.append("## 热门就业单位（录用人数排名）")
            for company, cnt in top_companies:
                lines.append(f"- {company}：{cnt}人")
            lines.append("")

        if top_industries:
            lines.append("## 主要就业行业")
            for ind, cnt in top_industries:
                lines.append(f"- {ind}：{cnt}人")
            lines.append("")

        if top_positions:
            lines.append("## 常见岗位")
            for pos, cnt in top_positions:
                lines.append(f"- {pos}：{cnt}人")
            lines.append("")

        if top_locations:
            lines.append("## 主要就业地区")
            for loc, cnt in top_locations:
                lines.append(f"- {loc}：{cnt}人")
            lines.append("")

        if grad_schools:
            lines.append("## 升学院校（考研/专升本）")
            for school, cnt in grad_schools:
                lines.append(f"- {school}：{cnt}人")
            lines.append("")

        chunks.append("\n".join(lines))

    return chunks


def gen_industry_chunks(records: List[dict], years: List[str]) -> List[str]:
    """每个主要行业生成薪酬分析 chunk"""
    chunks = []
    employed = [r for r in records if classify_record(r) == "就业" and r["industry"]]
    by_industry = defaultdict(list)
    for r in employed:
        by_industry[r["industry"]].append(r)

    # 只处理人数 >= 10 的行业
    for industry, recs in sorted(by_industry.items(), key=lambda x: -len(x[1])):
        if len(recs) < 10:
            continue

        sal = salary_stats([r["salary"] for r in recs])
        top_companies = top_items(Counter(r.get("company", "") for r in recs if r.get("company")))
        top_positions = top_items(Counter(r.get("position", "") for r in recs if r.get("position")))
        majors = top_items(Counter(r.get("major_full", "其他") for r in recs), 5)

        lines = [
            f"# {industry} 行业就业情况",
            f"数据年份：{'、'.join(years)}，样本：{len(recs)}人",
            "",
            "## 薪酬水平",
            f"- 平均薪酬：{sal['avg']} 元/月",
            f"- 中位数薪酬：{sal['median']} 元/月",
            f"- 薪酬区间：{sal['min']} ~ {sal['max']} 元/月",
            "",
        ]

        if top_companies:
            lines.append("## 代表性就业单位")
            for c, n in top_companies:
                lines.append(f"- {c}（{n}人）")
            lines.append("")

        if top_positions:
            lines.append("## 常见岗位")
            for p, n in top_positions:
                lines.append(f"- {p}（{n}人）")
            lines.append("")

        if majors:
            lines.append("## 主要来源专业")
            for maj, n in majors:
                lines.append(f"- {maj}（{n}人）")
            lines.append("")

        chunks.append("\n".join(lines))

    return chunks


def gen_graduate_vs_employment_chunk(records: List[dict], years: List[str]) -> List[str]:
    """生成升学 vs 就业对比 chunk"""
    total = len(records)
    by_major = defaultdict(list)
    for r in records:
        by_major[r["major_prefix"]].append(r)

    lines = [
        f"# 升学 vs 就业 对比分析",
        f"数据年份：{'、'.join(years)}，全校样本：{total}人",
        "",
        "## 全校总体",
    ]
    cats = Counter(classify_record(r) for r in records)
    for label in ["就业", "升学", "出国", "待定"]:
        cnt = cats.get(label, 0)
        lines.append(f"- {label}：{cnt}人（{cnt/total*100:.1f}%）")
    lines.append("")

    lines.append("## 各专业升学率排名")
    major_grad_rates = []
    for prefix, recs in by_major.items():
        full = MAJOR_MAP.get(prefix, prefix)
        cats_m = Counter(classify_record(r) for r in recs)
        grad_rate = cats_m.get("升学", 0) / len(recs) * 100
        major_grad_rates.append((full, grad_rate, cats_m.get("升学", 0), len(recs)))

    for full, rate, grad_cnt, total_m in sorted(major_grad_rates, key=lambda x: -x[1]):
        lines.append(f"- {full}：升学率 {rate:.1f}%（{grad_cnt}/{total_m}人）")
    lines.append("")

    lines.append("## 升学与就业薪酬差异")
    employed_sal = salary_stats([r["salary"] for r in records if classify_record(r) == "就业" and r["salary"]])
    if employed_sal["count"] > 0:
        lines.append(f"- 直接就业平均薪酬：{employed_sal['avg']} 元/月")
        lines.append(f"- 直接就业薪酬中位数：{employed_sal['median']} 元/月")
        lines.append("- 升学可提升学历竞争力，通常对应更高起薪，但需额外 2-3 年时间成本")
    lines.append("")

    return ["\n".join(lines)]


def gen_top_companies_chunk(records: List[dict], years: List[str]) -> List[str]:
    """生成热门公司推荐 chunk"""
    employed = [r for r in records if classify_record(r) == "就业" and r["company"]]
    by_company = defaultdict(list)
    for r in employed:
        by_company[r["company"]].append(r)

    lines = [
        f"# 热门就业单位推荐",
        f"数据年份：{'、'.join(years)}，录用人数 Top 30",
        "",
    ]

    sorted_companies = sorted(by_company.items(), key=lambda x: -len(x[1]))[:30]
    for company, recs in sorted_companies:
        sal = salary_stats([r["salary"] for r in recs])
        industries = Counter(r.get("industry", "") for r in recs if r.get("industry"))
        positions = Counter(r.get("position", "") for r in recs if r.get("position"))
        majors = Counter(r.get("major_full", "其他") for r in recs)

        top_ind = industries.most_common(1)[0][0] if industries else "未知"
        top_pos = positions.most_common(1)[0][0] if positions else "未知"
        top_maj = "、".join(m for m, _ in majors.most_common(3))

        sal_str = f"平均薪酬 {sal['avg']} 元" if sal["count"] > 0 else "薪酬未知"

        lines.append(f"## {company}（录用 {len(recs)} 人）")
        lines.append(f"- 所属行业：{top_ind}")
        lines.append(f"- 主要岗位：{top_pos}")
        lines.append(f"- {sal_str}")
        lines.append(f"- 主要招录专业：{top_maj}")
        lines.append("")

    return ["\n".join(lines)]


def gen_yearly_comparison_chunk(all_records_by_year: Dict[str, List[dict]]) -> List[str]:
    """多年数据对比 chunk"""
    if len(all_records_by_year) < 2:
        return []
    lines = ["# 历年就业数据对比", ""]
    for year in sorted(all_records_by_year.keys()):
        recs = all_records_by_year[year]
        cats = Counter(classify_record(r) for r in recs)
        employed = [r for r in recs if classify_record(r) == "就业"]
        sal = salary_stats([r["salary"] for r in employed])

        lines.append(f"## {year}年（共 {len(recs)} 人）")
        lines.append(f"- 就业：{cats.get('就业', 0)}人（{cats.get('就业', 0)/len(recs)*100:.1f}%）")
        lines.append(f"- 升学：{cats.get('升学', 0)}人（{cats.get('升学', 0)/len(recs)*100:.1f}%）")
        if sal["count"] > 0:
            lines.append(f"- 平均薪酬：{sal['avg']} 元/月（中位数 {sal['median']}）")
        lines.append("")

    return ["\n".join(lines)]

# ─────────────────────────────────────────────
# 把文本 chunk 转成 Document
# ─────────────────────────────────────────────
def texts_to_docs(chunks: List[str], source: str) -> List[Document]:
    docs = []
    for i, text in enumerate(chunks):
        doc_id = hashlib.md5(f"{source}_{i}_{text[:50]}".encode()).hexdigest()
        docs.append(Document(
            doc_id=doc_id,
            content=text,
            source=source,
            chunk_index=i,
            metadata={"type": "career_stats"},
        ))
    return docs

# ─────────────────────────────────────────────
# 主建库流程
# ─────────────────────────────────────────────
def build_career_knowledge_base(excel_files: List[str], show_chunks: bool = False):
    console.print(Panel(
        f"[bold green] 开始构建就业指导知识库 [/bold green]\n"
        f"文件：{', '.join(Path(f).name for f in excel_files)}"
    ))

    # Step 1: 读取所有 Excel
    console.print("\n[bold cyan]Step 1/4 读取 Excel 数据 [/bold cyan]")
    all_records = []
    records_by_year = {}

    for path in excel_files:
        year = re.search(r"(20\d{2})", Path(path).name)
        year = year.group(1) if year else Path(path).stem
        recs = load_excel(path, year)
        records_by_year[year] = recs
        all_records.extend(recs)

        cats = Counter(classify_record(r) for r in recs)
        console.print(
            f" [green]✓[/green] {Path(path).name} {year}年 {len(recs)}条记录 "
            f"（就业:{cats.get('就业', 0)} 升学:{cats.get('升学', 0)} 待定:{cats.get('待定', 0)}）"
        )

    years = sorted(records_by_year.keys())
    console.print(f"\n 合计 [cyan]{len(all_records)}[/cyan] 条记录，涵盖 {len(years)} 个年份")

    # Step 2: 生成统计摘要 chunk
    console.print("\n[bold cyan]Step 2/4 生成统计摘要 [/bold cyan]")
    all_chunks = []

    major_chunks = gen_major_chunks(all_records, years)
    all_chunks.extend(major_chunks)
    console.print(f" 专业就业全景：[cyan]{len(major_chunks)}[/cyan] 个 chunk（每专业一个）")

    industry_chunks = gen_industry_chunks(all_records, years)
    all_chunks.extend(industry_chunks)
    console.print(f" 行业薪酬分析：[cyan]{len(industry_chunks)}[/cyan] 个 chunk")

    grad_chunks = gen_graduate_vs_employment_chunk(all_records, years)
    all_chunks.extend(grad_chunks)
    console.print(f" 升学 vs 就业对比：[cyan]{len(grad_chunks)}[/cyan] 个 chunk")

    company_chunks = gen_top_companies_chunk(all_records, years)
    all_chunks.extend(company_chunks)
    console.print(f" 热门公司推荐：[cyan]{len(company_chunks)}[/cyan] 个 chunk")

    yearly_chunks = gen_yearly_comparison_chunk(records_by_year)
    all_chunks.extend(yearly_chunks)
    if yearly_chunks:
        console.print(f" 历年对比：[cyan]{len(yearly_chunks)}[/cyan] 个 chunk")

    console.print(f"\n 总计 [bold cyan]{len(all_chunks)}[/bold cyan] 个 chunk")

    # Step 3: chunk 预览
    if show_chunks:
        table = Table(show_lines=True)
        table.add_column("序号", width=4)
        table.add_column("内容预览", width=90)
        for i, chunk in enumerate(all_chunks):
            preview = chunk[:100].replace("\n", " ") + "..."
            table.add_row(str(i+1), preview)
        console.print(table)

    # Step 4: 转 Document + 建索引
    console.print("\n[bold cyan]Step 3/4 建立向量索引 + BM25 索引 [/bold cyan]")
    all_docs = texts_to_docs(all_chunks, source="career_stats")
    os.makedirs(CAREER_CONFIG.storage_dir, exist_ok=True)
    retriever = HybridRetriever(CAREER_CONFIG)
    retriever.build(all_docs)
    retriever.save()

    # Step 5: 保存
    console.print("\n[bold cyan]Step 4/4 保存文档块 [/bold cyan]")
    docs_path = os.path.join(CAREER_CONFIG.storage_dir, "docs.json")
    DocumentProcessor(CAREER_CONFIG).save_docs(all_docs, docs_path)

    # 同时保存原始统计数据供 pipeline 直接查询
    import json
    stats_path = os.path.join(CAREER_CONFIG.storage_dir, "stats_chunks.json")
    with open(stats_path, "w", encoding="utf-8") as f:
        json.dump({"years": years, "chunks": all_chunks}, f, ensure_ascii=False, indent=2)

    console.print(Panel(
        f"[bold green] 就业知识库构建完成！ [/bold green]\n\n"
        f" 数据年份：{'、'.join(years)}\n"
        f" 记录条数：{len(all_records)} 条\n"
        f" 知识块数：{len(all_chunks)} 个\n"
        f" 存储目录：{CAREER_CONFIG.storage_dir}/\n\n"
        f"[yellow]下一步：[/yellow]\n"
        f" python career_main.py chat",
        border_style="green"
    ))


def verify():
    docs_path = os.path.join(CAREER_CONFIG.storage_dir, "docs.json")
    console.print("\n[bold cyan] 验证就业知识库 [/bold cyan]")
    all_ok = True
    for label, path in [
        ("docs.json", docs_path),
        ("faiss.index", CAREER_CONFIG.embedding.index_path),
        ("bm25.pkl", CAREER_CONFIG.bm25.index_path),
    ]:
        if os.path.exists(path):
            size = os.path.getsize(path) / 1024
            console.print(f" [green]✓[/green] {label} ({size:.1f} KB)")
        else:
            console.print(f" [red]✗ 缺少 {label}[/red]")
            all_ok = False

    if all_ok:
        docs = DocumentProcessor(CAREER_CONFIG).load_docs(docs_path)
        console.print(f"\n 共 [cyan]{len(docs)}[/cyan] 个 chunk，知识库正常")
    else:
        console.print("\n[red]知识库不完整，请重新运行 build[/red]")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="就业知识库建库工具")
    subparsers = parser.add_subparsers(dest="command")

    build_p = subparsers.add_parser("build", help="从 Excel 建立就业知识库")
    build_p.add_argument("--files", nargs="+", required=True,
                         help="Excel 文件路径，支持多个，如 2023.xlsx 2024.xlsx 2025.xlsx")
    build_p.add_argument("--show", action="store_true", help="打印 chunk 预览")

    subparsers.add_parser("verify", help="验证知识库")

    args = parser.parse_args()

    if args.command == "build":
        build_career_knowledge_base(args.files, show_chunks=args.show)
    elif args.command == "verify":
        verify()
    else:
        parser.print_help()